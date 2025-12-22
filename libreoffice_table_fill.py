"""Extract and display address rows from a spreadsheet file.

This version reads the workbook directly with :mod:`openpyxl`, avoiding the
LibreOffice UNO bridge so it can run in a standard Python environment. The
input file location is still hard-coded, but the parsing logic remains the
same.
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List
from openpyxl import load_workbook

# Hard-coded path to the source spreadsheet
CALC_PATH = Path("data/source.xlsx")


@dataclass
class AddressEntry:
    """Normalized address data from a Calc row."""

    family_name: str
    family_info: str
    address_line_1: str
    address_apt_line: str
    address_final_line: str

    @property
    def formatted_text(self) -> str:
        lines: List[str] = []
        header = f"{self.family_info} {self.family_name}".strip()
        if header:
            lines.append(header)
        if self.address_line_1:
            lines.append(self.address_line_1)
        if self.address_apt_line:
            lines.append(self.address_apt_line)
        if self.address_final_line:
            lines.append(self.address_final_line)
        return "\n".join(lines)


def _safe_cell_value(value) -> str:
    """Normalize workbook cell values into stripped strings."""

    if value is None:
        return ""
    return str(value).strip()


def _normalize_family_info(raw: str, family_name: str) -> str:
    value = raw.strip()
    if value.lower() == "family":
        value = "Family"
    elif "(" in value and ")" in value:
        match = re.search(r"\(([^)]*)\)", value)
        if match:
            value = match.group(1).strip()
    value = re.sub(r"[^\w\s]", "", value).strip()
    if not value:
        value = "Family" if family_name else ""
    return value


def _parse_address_parts(column_c: str) -> tuple[str, str, str]:
    segments = [segment.strip() for segment in column_c.split(",")]
    address_line_1 = segments[0] if segments else ""
    middle = segments[1] if len(segments) > 1 else ""
    remainder = ",".join(segments[2:]).strip() if len(segments) > 2 else ""

    address_apt_line = ""
    address_final_line = remainder

    if middle:
        if "apt" in middle.lower():
            address_apt_line = middle
        else:
            address_final_line = " ".join(filter(None, [middle, remainder])).strip()
    return address_line_1, address_apt_line, address_final_line


def _load_first_worksheet(path: Path):
    workbook = load_workbook(path, data_only=True)
    return workbook.worksheets[0]


def extract_entries(worksheet) -> List[AddressEntry]:
    entries: List[AddressEntry] = []
    empty_rows = 0

    for row in worksheet.iter_rows(values_only=True):
        col_a = _safe_cell_value(row[0]) if len(row) > 0 else ""
        col_b = _safe_cell_value(row[1]) if len(row) > 1 else ""
        col_c = _safe_cell_value(row[2]) if len(row) > 2 else ""

        if not col_a and not col_b and not col_c:
            empty_rows += 1
            if empty_rows >= 5:
                break
            continue
        empty_rows = 0

        if col_a.upper() != "TRUE":
            continue

        words = col_b.split()
        if not words:
            continue
        family_name = words[0]
        family_info_raw = " ".join(words[1:])
        address_line_1, address_apt_line, address_final_line = _parse_address_parts(col_c)

        family_info = _normalize_family_info(family_info_raw, family_name)

        entries.append(
            AddressEntry(
                family_name=family_name,
                family_info=family_info,
                address_line_1=address_line_1,
                address_apt_line=address_apt_line,
                address_final_line=address_final_line,
            )
        )
    return entries


def _format_entries(entries: Iterable[AddressEntry]) -> str:
    """Combine formatted address blocks for easy copying."""

    return "\n\n".join(entry.formatted_text for entry in entries)


def main() -> None:
    worksheet = _load_first_worksheet(CALC_PATH)
    entries = extract_entries(worksheet)

    print(_format_entries(entries))


if __name__ == "__main__":
    main()
